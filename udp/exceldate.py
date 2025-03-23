from openpyxl import load_workbook


def excel_to_dict(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    result = {}

    # 从第二行开始遍历（跳过标题行）
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            # 跳过空行或列不足的行
            if len(row) < 2 or not row[0]:  # 检查键列是否为空
                continue

            # 第一列作为键，第二列作为值
            key = row[0]
            value = row[1]

            # 处理重复键的情况
            if key in result:
                print(f"警告：第 {i} 行键 '{key}' 已存在（第一列重复），将覆盖旧值")

            result[key] = value

        except Exception as e:
            print(f"第 {i} 行处理失败：{str(e)}")

    wb.close()
    return result


# 使用示例
data = excel_to_dict("C:/Users/born/Desktop/记事本/机房设备清单（正式版）.xlsx")
print("生成的字典：", data)