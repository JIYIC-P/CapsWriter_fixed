from openpyxl import load_workbook
from difflib import get_close_matches
class excel:
    data = None
    def read(file_path):
        workbook = load_workbook(file_path)
        表格 = {}
        上一个 = ""
        for sheet in workbook.worksheets:
            # 从第二行开始遍历（跳过标题行）
            if sheet.title.endswith('组机柜') or sheet.title.endswith('组设备'):
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # 跳过空行或列不足的行
                        if not row[1] == None:  #机柜编号列不为空
                            # 第一列作为键，第二列作为值
                            机柜编号 = row[1]
                            if 机柜编号 in 表格:
                                print(f"警告：第 {i-1} 行键 '{机柜编号}' 已存在（第一列重复），将覆盖旧值")
                            
                            if row[2] == None:
                                系统名称 = 上一个
                            else:
                                系统名称 = row[2]
                                上一个 = 系统名称
                            表格[机柜编号]=系统名称
                            
                    except Exception as e:
                        print(f"第 {i-1} 行处理失败：{str(e)}")
        workbook.close()
        return 表格

    def fix_excel(file_path):
        # 加载工作簿
        workbook = load_workbook(file_path)
        # 遍历每个工作表
        for sheet in workbook.worksheets:
            if sheet.title.endswith('组机柜') or sheet.title.endswith('组设备'):
                # 从第二行开始遍历（跳过标题行）
                rows_to_delete = []
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    # 检查第二列是否为空且第一列不为空
                    if row[1] is None and row[0] is not None:
                        rows_to_delete.append(i)
                    rowd = sorted(rows_to_delete, reverse=True)  # 确保大的行数首先删除
                    for r in rowd:                     # rowd格式如：[1,3,5],表示要删除第1、3、5共三行。
                        sheet.delete_rows(r)
        # 保存修改后的文件
        workbook.save(file_path)
        # 关闭工作簿
        workbook.close()

    def where(name: str, sheets_dict: dict, cutoff=0.5):
        matches = []
        for key, value in sheets_dict.items():
            # 检查 name 是否和 value 相似
            close_matches = get_close_matches(name, [value], n=1, cutoff=cutoff)
            if close_matches:
                matches.append(key)
                matches.append("和")  # 在每个匹配项后添加"和"
        if matches:  # 如果找到匹配项
            matches.pop()
            return matches
        
    def what(add, sheets_dict: dict):
        matche = sheets_dict.get(add)
        if matche is None:
            return []
        else :
            return matche
        
    def place(add: str, name: str):
        file_path = "C:\\Users\\14676\\Desktop\\副本机房设备清单（正式版）.xlsx"
        workbook = load_workbook(file_path)
        try:
            # 1. 找到匹配 add[0]（如 "A"、"B"）的工作表
            target_sheet = None
            for sheet in workbook.worksheets:
                if add[0] in sheet.title:
                    target_sheet = sheet
                    break
            if not target_sheet:
                print(f"未找到匹配 '{add[0]}' 的工作表！")
                return
            id = target_sheet.cell(row=target_sheet.max_row, column=1).value+1
            target_sheet.append([id, add, name])
            workbook.save(file_path)
            print(f"成功添加：序号={id}, 位置={add}, 设备={name}")
        except Exception as e:
            print(f"发生错误：{e}")
        finally:
            workbook.close()  # 确保关闭工作簿



if __name__ == "__main__" :
    path = "C:\\Users\\14676\\Desktop\\副本机房设备清单（正式版）.xlsx"
    excel.fix_excel(path)
    data = excel.read(path)
    print(data)
    #place("A10-10", "测试物品")
    excel.where("H3CX10516G三存储",data)
